import requests
from bs4 import BeautifulSoup
import openpyxl

# 주식회사 코드
codes = ['005930', '036570', '293490', '225570', '194480']

# 엑셀에 저장하기
wb = openpyxl.Workbook()
# ws = wb.create_sheet('오늘의 주식')
ws = wb.active  # 현재 활성화된 시트 선택 (ex:기본 시트)

ws['A1'] = '종목'
ws['B1'] = '현재가'

# 주식 정보 크롤링
for idx, code in enumerate(codes):
    response = requests.get(
        f'https://finance.naver.com/item/sise.naver?code={code}')
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')
    company_name = soup.select_one(
        '#middle > div.h_company > div.wrap_company > h2 > a').text
    stock_price = soup.select_one('#_nowVal').text
    print(f'{company_name}: {stock_price}')

    # 엑셀 파일에 저장
    ws[f'A{idx+2}'] = company_name
    ws[f'B{idx+2}'] = stock_price


wb.save(r'C:\SourceCode\crawling_practice\excel\주식정보_data.xlsx')
