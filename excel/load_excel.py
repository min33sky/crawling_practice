import openpyxl

excel_path = r'C:\SourceCode\crawling_practice\excel\참가자_data.xlsx'

# 1) 엑셀 불러오기
wb = openpyxl.load_workbook(excel_path)

# 2) 엑셀 시트 선택
ws = wb['오징어게임']

# 3) 데이터 수정하기
ws['A3'] = 456
ws['B3'] = '드록바'

# 4) 엑셀 저장하기
wb.save(excel_path)
