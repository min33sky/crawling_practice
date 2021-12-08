import openpyxl

# 1) 엑셀 만들기
wb = openpyxl.Workbook()

# 2) 엑셀 워크시트 만들기
ws = wb.create_sheet('오징어게임')

# 3) 데이터 추가하기
ws['A1'] = '참가번호'
ws['B1'] = '성명'

ws['A2'] = 1
ws['B2'] = '오일남'

# 4) 엑셀 저장하기 (r-string을 사용하면 row string 처리가 되어서 이스케이프 문자를 사용 안해도 됨)
wb.save(r'C:\SourceCode\crawling_practice\excel\참가자_data.xlsx')
